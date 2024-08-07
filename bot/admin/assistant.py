# -*- coding: UTF-8 -*-

from aiogram.utils.deep_linking import create_start_link
from aiogram.types import FSInputFile
from typing import Optional, Tuple
from aiogram import Bot


import openpyxl
import zipfile
import logging
import qrcode
import json
import os


from dotenv import set_key


class AdminOperations:
    def __init__(self) -> None:
        self.auth_client = None
    
    
    async def get_manager_password(self) -> str:
        """
        Получение пароля админа бота
        """
        return os.environ.get('MANAGER_PASSWORD_HASH')
    
    
    def get_attempts_enter_wrong_password(self) -> str:
        """
        Получение количества оставшихся попыток ввода пароля
        """
        return os.environ.get('MAX_ATTEMPTS')
    
    
    async def update_attempts_enter_wrong_password(self, remaining_attempts: dict) -> str:
        """
        Обновление количества оставшихся попыток ввода пароля
        """
        set_key('docker-compose.env', 'MAX_ATTEMPTS', str(remaining_attempts))
        return os.environ.get('MAX_ATTEMPTS')
    
    
    async def update_list_admin_users(self, list_users: dict) -> str:
        """
        Обновление количества оставшихся попыток ввода пароля
        """
        try:
            list_users = json.dumps(list_users)
            set_key('docker-compose.env', 'ADMIN_USERS', str(list_users))
        except Exception as e:
            logging.error(f"Error during update_list_admin_users: {e}")
    
    
    async def update_list_banned_users(self, list_users: dict) -> str:
        """
        Обновление количества оставшихся попыток ввода пароля
        """
        try:
            list_users = json.dumps(list_users)
            set_key('docker-compose.env', 'BANNED_USERS', str(list_users))
        except Exception as e:
            logging.error(f"Error during update_list_banned_users: {e}")
            
    
    async def download_and_save_exel_document(self, file_id: str, file_name: str, bot: Bot) -> None:
        try:
            file = await bot.get_file(file_id)
            file_path = file.file_path
            downloaded_file = await bot.download_file(file_path)

            # Сохраняем файл на сервере
            path_table = f'./downloads/{file_name}'
            with open(path_table, 'wb') as new_file:
                new_file.write(downloaded_file.read())
                
            return path_table
        except Exception as e:
            logging.error(f"Error during download file: {file_path}, Code: {e}") 
            
    
    async def generate_zip_qr_codes(self, uploaded_data: str, bot: Bot) -> None:
        """
        Генерация фото с QR-кодами
        """
        for row in uploaded_data:
            qr = qrcode.QRCode(
                version=1,
                error_correction=qrcode.constants.ERROR_CORRECT_L,
                box_size=10,
                border=4,
            )
            date = row[0][1]
            date = date.replace('.', '')
            stroke = date + "-" + str(row[0][2]) + "-" +  str(row[0][3])
            #await message.answer(stroke)
            link = await create_start_link(bot, stroke)
            link = link.replace('-', '=')
            qr.add_data(link)
            qr.make(fit=True)
            img = qr.make_image(fill_color="black", back_color="white")
            name_qr_code = 'qr_code' + str(row[0][2]) + '.png'
            img.save(f'./photos/{name_qr_code}')
    
        zip_filename = 'qr_codes.zip'
        qr_codes_folder = './photos'
        
        await self.create_zip(zip_filename, qr_codes_folder, "w")
        zip_to_send = FSInputFile(zip_filename)
        
        items = os.listdir('photos')
        for photo in items:
            os.remove('photos' +'/' +photo)
        
        return zip_to_send, zip_filename
                

    async def parse_callback_data(self, data: str) -> Tuple[str, Optional[str], Optional[str]]:
        """
        Получение данных из строки callback.data
        """
        if ',' in data:
            parts = data.split(',')
            action = parts[0]
            user_id = parts[1] if len(parts) > 1 else None
            user_tg_addr = ','.join(parts[2:]) if len(parts) > 2 else None
            return action, user_id, user_tg_addr
        else:
            return data, None, None
    
    
    async def create_zip(self, arch, folder_list, mode):
        """
        Генерация архива с QR-кодами
        """
        num = 0
        num_ignore = 0
        z = zipfile.ZipFile(arch, mode, zipfile.ZIP_DEFLATED, True)
        # Получаем папки из списка папок.
        items = os.listdir(folder_list)
        for file in items:
                    z.write(folder_list + '/' + file)
                    print(num, folder_list + '/' + file)
                    num += 1
        z.close()
        logging.info("------------------------------")
        logging.info(f"Добавлено: {num}")
        

    async def create_database_data_exel_document(self, type_data: str, user_data: dict, scanned_data: dict, unscanned_data: dict, user_id: str) -> str:
        """
        Создание Exel документов с данными для выгрузки
        """
        sells_ids = ['A1', 'B1', 'C1', 'D1', 'E1', 'F1', 'G1']
        user_data_sells_names = ['ID', 'ID Телеграмм аккаунта', 'Никнейм', 'Имя', 'Номер телефона', 'Дата создания аккаунта', 'Кол-во сканирований']
        scanned_product_sells_names = ['ID', 'Дата поставки', 'Номер упаковки', 'Идентификатор', 'Сканировщик', 'Дата сканирования']
        unscanned_product_sells_names = ['ID', 'Дата поставки', 'Номер упаковки', 'Идентификатор']
        
        workbook = openpyxl.Workbook()
        
        if type_data == 'user_data':
            sheet = workbook.active
            sheet.title = 'Пользователи'
            for sell_id, sell_name in zip(sells_ids, user_data_sells_names):
                sheet[sell_id] = sell_name
            
            for row, row_data in enumerate(user_data, start=2):
                sheet.cell(row=row, column=1, value=row_data[0])
                sheet.cell(row=row, column=2, value=row_data[1])
                sheet.cell(row=row, column=3, value=row_data[2])
                sheet.cell(row=row, column=4, value=row_data[3])
                sheet.cell(row=row, column=5, value=row_data[4])
                sheet.cell(row=row, column=6, value=row_data[5].strftime('%d.%m.%Y %H:%M'))
                sheet.cell(row=row, column=7, value=row_data[6])
            
            exel_document_name = f"Таблица пользователей{user_id}.xlsx"
            workbook.save(exel_document_name)

            return exel_document_name
        elif type_data == 'scanned_product':
            sheet = workbook.active
            sheet.title = 'Отсканированные продукты'
            for sell_id, sell_name in zip(sells_ids[:6], scanned_product_sells_names):
                sheet[sell_id] = sell_name
            
            for row, row_data in enumerate(scanned_data, start=2):
                sheet.cell(row=row, column=1, value=row_data[0])
                sheet.cell(row=row, column=2, value=row_data[1])
                sheet.cell(row=row, column=3, value=row_data[2])
                sheet.cell(row=row, column=4, value=row_data[3])
                sheet.cell(row=row, column=5, value=row_data[4])
                sheet.cell(row=row, column=6, value=row_data[5].strftime('%d.%m.%Y %H:%M'))

            zip_name = f"Таблица отсканированного товара{user_id}.xlsx"
            workbook.save(exel_document_name)

            return exel_document_name
        
        elif type_data == 'unscanned_product':
            sheet = workbook.active
            sheet.title = 'Не отсканированные продукты'
            for sell_id, sell_name in zip(sells_ids[:4], unscanned_product_sells_names):
                sheet[sell_id] = sell_name
            
            for row, row_data in enumerate(unscanned_data, start=2):
                sheet.cell(row=row, column=1, value=row_data[0])
                sheet.cell(row=row, column=2, value=row_data[1])
                sheet.cell(row=row, column=3, value=row_data[2])
                sheet.cell(row=row, column=4, value=row_data[3])
            
            exel_document_name = f"Таблица не отсканированного товара{user_id}.xlsx"
            workbook.save(exel_document_name)

            return exel_document_name
        
        elif type_data == 'full_db_export':
            sheet = workbook.active
            users_sheet = workbook.create_sheet('users')
            scanned_product_sheet = workbook.create_sheet('scanned_product')
            unscanned_product_sheet = workbook.create_sheet('unscanned_product')
            
            for sell_id, sell_name in zip(sells_ids, user_data_sells_names):
                users_sheet[sell_id] = sell_name
                
            for sell_id, sell_name in zip(sells_ids[:6], scanned_product_sells_names):
                scanned_product_sheet[sell_id] = sell_name
            
            for sell_id, sell_name in zip(sells_ids[:4], unscanned_product_sells_names):
                unscanned_product_sheet[sell_id] = sell_name
            
            for row, row_data in enumerate(user_data, start=2):
                users_sheet.cell(row=row, column=1, value=row_data[0])
                users_sheet.cell(row=row, column=2, value=row_data[1])
                users_sheet.cell(row=row, column=3, value=row_data[2])
                users_sheet.cell(row=row, column=4, value=row_data[3])
                users_sheet.cell(row=row, column=5, value=row_data[4])
                users_sheet.cell(row=row, column=6, value=row_data[5].strftime('%d.%m.%Y %H:%M'))
                users_sheet.cell(row=row, column=7, value=row_data[6])
                
            for row, row_data in enumerate(scanned_data, start=2):
                scanned_product_sheet.cell(row=row, column=1, value=row_data[0])
                scanned_product_sheet.cell(row=row, column=2, value=row_data[1])
                scanned_product_sheet.cell(row=row, column=3, value=row_data[2])
                scanned_product_sheet.cell(row=row, column=4, value=row_data[3])
                scanned_product_sheet.cell(row=row, column=5, value=row_data[4])
                scanned_product_sheet.cell(row=row, column=6, value=row_data[5].strftime('%d.%m.%Y %H:%M'))
            
            for row, row_data in enumerate(unscanned_data, start=2):
                unscanned_product_sheet.cell(row=row, column=1, value=row_data[0])
                unscanned_product_sheet.cell(row=row, column=2, value=row_data[1])
                unscanned_product_sheet.cell(row=row, column=3, value=row_data[2])
                unscanned_product_sheet.cell(row=row, column=4, value=row_data[3])
            
            exel_document_name = f"Выгрузка базы данных{user_id}.xlsx"
            workbook.save(exel_document_name)

            return exel_document_name
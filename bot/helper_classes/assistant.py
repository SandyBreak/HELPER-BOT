# -*- coding: UTF-8 -*-

import openpyxl
import hashlib
import zipfile
import random
import string
import os


from data_storage.dataclasses import ConnectionData, UserData, ProductData
from exeptions import *


class MinorOperations:
    def __init__(self) -> None:
        pass
    
    
    async def get_tg_token(self) -> str:
        return os.environ.get('TELEGRAM_TOKEN')
    
    
    def get_mongo_login(self) -> str:
        return os.environ.get('MONGO_INITDB_ROOT_USERNAME')
    
    
    def get_mongo_password(self) -> str:
        return os.environ.get('MONGO_INITDB_ROOT_PASSWORD')
    
    
    async def get_root_password(self) -> str:
        return os.environ.get('MARIADB_ROOT_PASSWORD')
    
    
    async def get_new_user_login(self) -> str:
        return os.environ.get('MARIADB_USER_NAME')
    
    
    async def get_new_user_password(self) -> str:
        return os.environ.get('MARIADB_USER_PASSWORD')
    
    
    async def get_manager_password(self) -> str:
        return os.environ.get('MANAGER_PASSWORD_HASH')
    
    
    async def get_attempts_enter_wrong_password(self) -> str:
        return os.environ.get('MAX_ATTEMPTS')
    
    async def fill_connection_data(self) -> dict:
        connect = ConnectionData(
            host=os.environ.get('MARIADB_HOST'),
            port=int(os.environ.get('MARIADB_PORT')),
            user=os.environ.get('MARIADB_USER_NAME'),
            password=os.environ.get('MARIADB_USER_PASSWORD'),
            db_name=os.environ.get('MARIADB_DB_NAME')
        ) 
        return connect
    
    
    async def fill_user_data(self, user_id: str, nick: str, full_name: str, phone_number: str) -> dict:
        user_data = UserData(
            id_tg=user_id,
            nickname='@' + nick,
            full_name=full_name,
            phone=phone_number
        ) 
        return user_data
    
    
    async def fill_product_data(self, args: dict) -> dict:
        try:
            params = args.split('=')
            refactored_date = f"{params[0][:2]}.{params[0][2:4]}.{params[0][4:]}"
            
            product_data = ProductData(
                date_of_man=refactored_date,
                number_packgage=params[1],
                identifier= params[2]
            )
            return product_data
        
        except Exception as e:
            raise WrongDataQrCode
        
    
    async def generate_identifier(self) -> str:
        characters = string.ascii_lowercase + string.digits
        random_string = ''.join(random.choice(characters) for _ in range(8))

        hashed_string = hashlib.sha256(random_string.encode()).hexdigest()
        return hashed_string[28:36]
    
    
    async def create_user_data_exel_document(self, user_data:dict, user_id: str) -> str:
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        sheet['A1'] = 'ID'
        sheet['B1'] = 'ID Телеграмм аккаунта'
        sheet['C1'] = 'Никнейм'
        sheet['D1'] = 'Имя'
        sheet['E1'] = 'Номер телефона'
        sheet['F1'] = 'Дата создания аккаунта'
        sheet['G1'] = 'Кол-во сканирований'
        
        for row, row_data in enumerate(user_data, start=2):
            sheet.cell(row=row, column=1, value=row_data[0])
            sheet.cell(row=row, column=2, value=row_data[1])
            sheet.cell(row=row, column=3, value=row_data[2])
            sheet.cell(row=row, column=4, value=row_data[3])
            sheet.cell(row=row, column=5, value=row_data[4])
            sheet.cell(row=row, column=6, value=row_data[5].strftime('%d.%m.%Y %H:%M'))
            sheet.cell(row=row, column=7, value=row_data[6])
        
        zip_name = f"Таблица пользователей{user_id}.xlsx"
        workbook.save(zip_name)
        
        return zip_name
    
    
    async def create_scanned_data_exel_document(self, scanned_data:dict, user_id: str) -> str:
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        sheet['A1'] = 'ID'
        sheet['B1'] = 'Дата поставки'
        sheet['C1'] = 'Номер упаковки'
        sheet['D1'] = 'Идентификатор'
        sheet['E1'] = 'Сканировщик'
        sheet['F1'] = 'Дата сканирования'
        
        for row, row_data in enumerate(scanned_data, start=2):
            sheet.cell(row=row, column=1, value=row_data[0])
            sheet.cell(row=row, column=2, value=row_data[1])
            sheet.cell(row=row, column=3, value=row_data[2])
            sheet.cell(row=row, column=4, value=row_data[3])
            sheet.cell(row=row, column=5, value=row_data[4])
            sheet.cell(row=row, column=6, value=row_data[5].strftime('%d.%m.%Y %H:%M'))
        
        zip_name = f"Таблица отсканированного товара{user_id}.xlsx"
        workbook.save(zip_name)
        
        return zip_name
    
    
    async def create_unscanned_data_exel_document(self, unscanned_data:dict, user_id: str) -> str:
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        sheet['A1'] = 'ID'
        sheet['B1'] = 'Дата поставки'
        sheet['C1'] = 'Номер упаковки'
        sheet['D1'] = 'Идентификатор'
        
        for row, row_data in enumerate(unscanned_data, start=2):
            sheet.cell(row=row, column=1, value=row_data[0])
            sheet.cell(row=row, column=2, value=row_data[1])
            sheet.cell(row=row, column=3, value=row_data[2])
            sheet.cell(row=row, column=4, value=row_data[3])

        zip_name = f"Таблица не отсканированного товара{user_id}.xlsx"
        workbook.save(zip_name)
        
        return zip_name
    
    
    async def create_full_bd_exel_document(self, user_data:dict, scanned_data:dict, unscanned_data:dict, user_id: str) -> str:
        workbook = openpyxl.Workbook()

        # Создаем три листа
        sheet1 = workbook.create_sheet('users')
        sheet2 = workbook.create_sheet('scanned_product')
        sheet3 = workbook.create_sheet('unscanned_product')
        
        sheet1['A1'] = 'ID'
        sheet1['B1'] = 'ID Телеграмм аккаунта'
        sheet1['C1'] = 'Никнейм'
        sheet1['D1'] = 'Имя'
        sheet1['E1'] = 'Номер телефона'
        sheet1['F1'] = 'Дата создания аккаунта'
        sheet1['G1'] = 'Кол-во сканирований'
        
        for row, row_data in enumerate(user_data, start=2):
            sheet1.cell(row=row, column=1, value=row_data[0])
            sheet1.cell(row=row, column=2, value=row_data[1])
            sheet1.cell(row=row, column=3, value=row_data[2])
            sheet1.cell(row=row, column=4, value=row_data[3])
            sheet1.cell(row=row, column=5, value=row_data[4])
            sheet1.cell(row=row, column=6, value=row_data[5].strftime('%d.%m.%Y %H:%M'))
            sheet1.cell(row=row, column=7, value=row_data[6])
        
        
        sheet2['A1'] = 'ID'
        sheet2['B1'] = 'Дата поставки'
        sheet2['C1'] = 'Номер упаковки'
        sheet2['D1'] = 'Идентификатор'
        sheet2['E1'] = 'Сканировщик'
        sheet2['F1'] = 'Дата сканирования'
        
        for row, row_data in enumerate(scanned_data, start=2):
            sheet2.cell(row=row, column=1, value=row_data[0])
            sheet2.cell(row=row, column=2, value=row_data[1])
            sheet2.cell(row=row, column=3, value=row_data[2])
            sheet2.cell(row=row, column=4, value=row_data[3])
            sheet2.cell(row=row, column=5, value=row_data[4])
            sheet2.cell(row=row, column=6, value=row_data[5].strftime('%d.%m.%Y %H:%M'))
        
        sheet3['A1'] = 'ID'
        sheet3['B1'] = 'Дата поставки'
        sheet3['C1'] = 'Номер упаковки'
        sheet3['D1'] = 'Идентификатор'
        
        for row, row_data in enumerate(unscanned_data, start=2):
            sheet3.cell(row=row, column=1, value=row_data[0])
            sheet3.cell(row=row, column=2, value=row_data[1])
            sheet3.cell(row=row, column=3, value=row_data[2])
            sheet3.cell(row=row, column=4, value=row_data[3])
        
        zip_name = f"Выгрузка базы данных{user_id}.xlsx"
        workbook.save(zip_name)
        
        return zip_name
    
    async def create_zip(self, arch, folder_list, mode):
        num = 0
        num_ignore = 0
        z = zipfile.ZipFile(arch, mode, zipfile.ZIP_DEFLATED, True)
        # Получаем папки из списка папок.
        items = os.listdir(folder_list)
        for file in items:
                    z.write(folder_list + '/' + file)
                    print(num, folder_list + '/' + file)
                    num += 1
        z.close()
        print("------------------------------")
        print("Добавлено: ", num)
        print("Проигнорировано: ", num_ignore)